import openpyxl
from typing import Any, Dict, List, Optional, Union
from utils.logger import global_logger
from utils.json_utils import safe_parse_json

class ExcelReader:
    JSON_FIELDS = ['测试body', '预期结果']
    
    def __init__(self, file_path: str, ignore_array_order: bool = False):
        self.file_path = file_path
        self.workbook = None
        self.logger = global_logger
        self.ignore_array_order = ignore_array_order
    
    def load_workbook(self) -> None:
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.logger.info(f"成功加载Excel文件: {self.file_path}")
        except Exception as e:
            self.logger.error(f"加载Excel文件失败: {e}")
            raise
    
    def get_sheet_data(self, sheet_name: str) -> List[Dict]:
        if not self.workbook:
            self.load_workbook()
        
        if sheet_name not in self.workbook.sheetnames:
            self.logger.warning(f"Sheet '{sheet_name}' 不存在")
            return []
        
        sheet = self.workbook[sheet_name]
        data = []
        
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            
            row_data = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_data[headers[i]] = value
            
            data.append(row_data)
        
        self.logger.info(f"从Sheet '{sheet_name}' 读取到 {len(data)} 条数据")
        return data
    
    def filter_executable_cases(self, data: List[Dict]) -> List[Dict]:
        executable_cases = []
        
        for case in data:
            is_executable = case.get('用例是否执行', 'FALSE')
            
            if isinstance(is_executable, str):
                is_executable = is_executable.upper().strip()
            
            if is_executable == True or is_executable == 'TRUE':
                for field in self.JSON_FIELDS:
                    if field in case and case[field]:
                        case[field] = safe_parse_json(case[field], strict=True)
                
                if 'ignore_array_order' not in case:
                    case['ignore_array_order'] = self.ignore_array_order
                
                executable_cases.append(case)
        
        self.logger.info(f"过滤后可执行的用例数: {len(executable_cases)}")
        return executable_cases
    
    def get_test_data(self, sheet_name: str) -> List[Dict]:
        data = self.get_sheet_data(sheet_name)
        return self.filter_executable_cases(data)
    
    def get_flow_sheets(self) -> List[str]:
        if not self.workbook:
            self.load_workbook()
        
        flow_sheets = []
        for sheet_name in self.workbook.sheetnames:
            if not sheet_name.startswith('-'):
                flow_sheets.append(sheet_name)
        
        self.logger.info(f"发现 {len(flow_sheets)} 个流程Sheet")
        return flow_sheets
    
    def get_flow_data(self, sheet_name: str) -> Dict:
        if not self.workbook:
            self.load_workbook()
        
        if sheet_name not in self.workbook.sheetnames:
            self.logger.warning(f"流程Sheet '{sheet_name}' 不存在")
            return {"flow_name": sheet_name, "steps": []}
        
        sheet = self.workbook[sheet_name]
        steps = []
        
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            
            row_data = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_data[headers[i]] = value
            
            if row_data.get("接口地址"):
                steps.append(row_data)
        
        self.logger.info(f"流程 '{sheet_name}' 包含 {len(steps)} 个步骤")
        return {
            "flow_name": sheet_name,
            "steps": steps
        }
    
    def get_all_flow_data(self) -> List[Dict]:
        flow_sheets = self.get_flow_sheets()
        all_flows = []
        
        for sheet_name in flow_sheets:
            flow_data = self.get_flow_data(sheet_name)
            if flow_data["steps"]:
                all_flows.append(flow_data)
        
        self.logger.info(f"共读取 {len(all_flows)} 个流程")
        return all_flows
    
    def close(self) -> None:
        if self.workbook:
            self.workbook.close()
            self.logger.info("关闭Excel文件")

def get_excel_data(file_path: str, sheet_name: str, ignore_array_order: bool = False) -> List[Dict]:
    reader = ExcelReader(file_path, ignore_array_order=ignore_array_order)
    try:
        data = reader.get_test_data(sheet_name)
        return data
    finally:
        reader.close()