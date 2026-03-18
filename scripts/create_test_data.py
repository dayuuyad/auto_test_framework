import openpyxl
from openpyxl import Workbook
import os

def create_api_test_data():
    excel_path = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'excel', 'api_test_data.xlsx')
    excel_path = os.path.abspath(excel_path)
    
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    
    wb = Workbook()
    
    ws_register = wb.active
    ws_register.title = "test_user_register"
    
    ws_register.append(["用例描述", "测试body", "用例是否执行"])
    
    ws_register.append([
        "正常注册用户-有效数据",
        '{"username":"testuser001","email":"testuser001@example.com","password":"Test@123456"}',
        "TRUE"
    ])
    
    ws_register.append([
        "注册用户-用户名已存在",
        '{"username":"admin","email":"admin@example.com","password":"Test@123456"}',
        "FALSE"
    ])
    
    ws_register.append([
        "注册用户-邮箱格式错误",
        '{"username":"testuser002","email":"invalid-email","password":"Test@123456"}',
        "TRUE"
    ])
    
    ws_register.append([
        "注册用户-密码长度不足",
        '{"username":"testuser003","email":"testuser003@example.com","password":"123"}',
        "TRUE"
    ])
    
    ws_login = wb.create_sheet("test_user_login")
    
    ws_login.append(["用例描述", "测试body", "用例是否执行"])
    
    ws_login.append([
        "正常登录-有效凭证",
        '{"email":"admin@example.com","password":"admin123"}',
        "TRUE"
    ])
    
    ws_login.append([
        "登录失败-错误密码",
        '{"email":"admin@example.com","password":"wrongpassword"}',
        "TRUE"
    ])
    
    ws_login.append([
        "登录失败-用户不存在",
        '{"email":"nonexist@example.com","password":"Test@123456"}',
        "TRUE"
    ])
    
    ws_login.append([
        "登录失败-邮箱格式错误",
        '{"email":"invalid-email","password":"Test@123456"}',
        "FALSE"
    ])
    
    ws_get_user = wb.create_sheet("test_get_user_info")
    
    ws_get_user.append(["用例描述", "测试body", "用例是否执行"])
    
    ws_get_user.append([
        "获取用户信息-有效用户ID",
        '{"user_id":1}',
        "TRUE"
    ])
    
    ws_get_user.append([
        "获取用户信息-用户不存在",
        '{"user_id":99999}',
        "TRUE"
    ])
    
    ws_get_user.append([
        "获取用户信息-无效用户ID",
        '{"user_id":-1}',
        "TRUE"
    ])
    
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(excel_path)
    print(f"Excel文件创建成功: {excel_path}")
    return excel_path

if __name__ == "__main__":
    create_api_test_data()